# qualifiers.py

import os
import logging
import pandas as pd
from rag import retrieve_answers_for_controls, load_faiss_index
from llm_analysis import call_ollama_api
from sentence_transformers import SentenceTransformer
import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def is_report_latest(df_chunks, model, index, top_k=3):
    """
    Determines if the SOC 2 Type 2 Report is the latest (published within last 12 months)
    by extracting the publication date from the retrieved text chunks.
    """
    query = "What is the publication date of the SOC 2 Type 2 Report?"
    query_emb = model.encode([query], show_progress_bar=False).astype('float32')

    distances, indices = index.search(query_emb, top_k)
    retrieved_answers = retrieve_answers_for_controls(
        pd.DataFrame([{'Control': query}]), model, index, df_chunks, top_k
    )

    current_date = datetime.datetime.now().date()

    prompt = f'''
    You are an expert in SOC 2 Type 2 compliance. Based solely on the following retrieved responses, determine the publication date of the SOC 2 Type 2 Report and assess whether it is the latest report.

    Retrieved Responses:
    {retrieved_answers.to_dict(orient='records')}

    Instructions:
    1. Extract the publication date of the SOC 2 Type 2 Report from the retrieved responses.
    2. Compare the extracted publication date with the current date ({current_date}).
    3. If the report was published within the last 12 months from the current date, it is considered the latest.
    4. Provide a clear and concise answer in the following exact format:
       - If the report is latest:
         "Yes. The SOC 2 Type 2 Report is the latest because it was published on [publication_date], which is within the last 12 months."
       - If the report is not latest:
         "No. The SOC 2 Type 2 Report is not the latest because it was published on [publication_date], which is more than 12 months ago."
    5. Do not include any additional information or commentary beyond the specified format.
    '''

    logging.debug("Prompt for 'is_report_latest': %s", prompt)
    response = call_ollama_api(prompt)
    return response


def are_trust_principles_covered(df_chunks, model, index, top_k=3):
    """
    Checks if the SOC 2 Type 2 Report explicitly covers Security, Availability, Confidentiality.
    """
    query = "Does the SOC 2 Type 2 Report cover the principles of Security, Availability, and Confidentiality?"
    query_emb = model.encode([query], show_progress_bar=False).astype('float32')

    distances, indices = index.search(query_emb, top_k)
    retrieved_answers = retrieve_answers_for_controls(
        pd.DataFrame([{'Control': query}]), model, index, df_chunks, top_k
    )

    prompt = f'''
    You are an expert in SOC 2 Type 2 compliance. Based solely on the following retrieved responses, determine whether the SOC 2 Type 2 Report covers all three Trust Principles: Security, Availability, and Confidentiality.

    Retrieved Responses:
    {retrieved_answers.to_dict(orient='records')}

    Instructions:
    1. Analyze the retrieved responses to identify mentions of the following Trust Principles:
       - Security
       - Availability
       - Confidentiality
    2. Determine if all three principles are explicitly covered in the SOC 2 Type 2 Report.
    3. Provide a clear and concise answer in the following exact format:
       - If all principles are covered:
         "Yes. The SOC 2 Type 2 Report covers all three Trust Principles (Security, Availability, Confidentiality) because [specific reasons]."
       - If any principle is not covered:
         "No. The SOC 2 Type 2 Report does not cover all three Trust Principles (Security, Availability, Confidentiality) because [specific reasons]."
    4. Ensure that the reasoning specifically addresses each principle mentioned or omitted.
    5. Do not include any additional information or commentary beyond the specified format.
    '''

    logging.debug("Prompt for 'are_trust_principles_covered': %s", prompt)
    response = call_ollama_api(prompt)
    return response


def is_audit_period_sufficient(df_chunks, model, index, top_k=3):
    """
    Checks if the audit period is at least 9 months by extracting start/end from the text.
    """
    query = "Does the SOC 2 Type 2 Report cover an audit period of at least 9 months?"
    query_emb = model.encode([query], show_progress_bar=False).astype('float32')

    distances, indices = index.search(query_emb, top_k)
    retrieved_answers = retrieve_answers_for_controls(
        pd.DataFrame([{'Control': query}]), model, index, df_chunks, top_k
    )

    prompt = f'''
    You are an expert in SOC 2 Type 2 compliance. Based solely on the following retrieved responses, determine whether the audit period covered in the SOC 2 Type 2 Report is at least 9 months.

    Retrieved Responses:
    {retrieved_answers.to_dict(orient='records')}

    Instructions:
    1. Extract the start and end dates of the audit period from the retrieved responses.
    2. Calculate the total duration of the audit period in months.
       - Consider partial months as full months for simplicity (e.g., from February 15 to May 14 is considered 3 months).
    3. If the audit period is 9 months or longer, it is considered sufficient.
    4. Provide a clear and concise answer in the following exact format:
       - If the audit period is sufficient:
         "Yes. The SOC 2 Type 2 Report covers an audit period of [duration] months, which meets the requirement of at least 9 months."
       - If the audit period is insufficient:
         "No. The SOC 2 Type 2 Report covers an audit period of [duration] months, which does not meet the requirement of at least 9 months."
    5. Do not include any additional information or commentary beyond the specified format.

    Example:
    Retrieved Responses:
    ["Response": "For the Period February 1, 2023 to May 31, 2023"]

    Analysis:
    - Start date: February 1, 2023
    - End date: May 31, 2023
    - Duration: 4 months

    Answer:
    "No. The SOC 2 Type 2 Report covers an audit period of 4 months, which does not meet the requirement of at least 9 months."
    '''

    logging.debug("Prompt for 'is_audit_period_sufficient': %s", prompt)
    response = call_ollama_api(prompt)

    # Attempt basic parse to ensure we have consistent results
    match = re.search(r'covers an audit period of (\d+) months', response)
    if match:
        duration = int(match.group(1))
        if duration >= 9:
            expected_response = f"Yes. The SOC 2 Type 2 Report covers an audit period of {duration} months, which meets the requirement of at least 9 months."
        else:
            expected_response = f"No. The SOC 2 Type 2 Report covers an audit period of {duration} months, which does not meet the requirement of at least 9 months."

        # If the LLM gave the exact expected response, good; otherwise we do a fallback
        if response.strip() == expected_response:
            return response
        else:
            logging.warning("LLM response does not match the exact expected format. Using the expected format.")
            return expected_response
    else:
        logging.error("Failed to parse the LLM response for audit period duration. Returning fallback.")
        return "No. The SOC 2 Type 2 Report does not provide a clear audit period duration."


def has_invalid_observations(df_chunks, model, index, top_k=3):
    """
    Checks if there are any observations in the independent auditor’s opinion that signify the report is invalid.
    """
    query = "Are there any observations in the independent auditor’s opinion that signify the SOC 2 Type 2 Report is invalid?"
    query_emb = model.encode([query], show_progress_bar=False).astype('float32')

    distances, indices = index.search(query_emb, top_k)
    retrieved_answers = retrieve_answers_for_controls(
        pd.DataFrame([{'Control': query}]), model, index, df_chunks, top_k
    )

    prompt = f'''
    You are an expert in SOC 2 Type 2 compliance. Based solely on the following retrieved responses, determine whether there are any observations in the independent auditor’s opinion that signify the SOC 2 Type 2 Report is invalid.

    Retrieved Responses:
    {retrieved_answers.to_dict(orient='records')}

    Instructions:
    1. Analyze the retrieved responses for any observations or remarks made by the independent auditor that could indicate the report is invalid.
    2. Provide a clear and concise answer in the following exact format:
       - If there are invalid observations:
         "Yes. The independent auditor’s opinion includes the following observations that signify the report is invalid: [list of observations]."
       - If there are no invalid observations:
         "No. There are no observations in the independent auditor’s opinion that signify the report is invalid."
    3. Do not include any additional information or commentary beyond the specified format.
    '''

    logging.debug("Prompt for 'has_invalid_observations': %s", prompt)
    response = call_ollama_api(prompt)
    return response


def is_report_qualified(df_chunks, model, index, top_k=3):
    """
    Checks if the SOC 2 Type 2 Report is a qualified report.
    """
    query = "Is the SOC 2 Type 2 Report a qualified report?"
    query_emb = model.encode([query], show_progress_bar=False).astype('float32')

    distances, indices = index.search(query_emb, top_k)
    retrieved_answers = retrieve_answers_for_controls(
        pd.DataFrame([{'Control': query}]), model, index, df_chunks, top_k
    )

    prompt = f'''
    You are an expert in SOC 2 Type 2 compliance. Based solely on the following retrieved responses, determine whether the SOC 2 Type 2 Report is a qualified report.

    Retrieved Responses:
    {retrieved_answers.to_dict(orient='records')}

    Instructions:
    1. Analyze the retrieved responses to determine if the SOC 2 Type 2 Report is a qualified report.
       - A "qualified" report indicates that there are reservations or issues with the report.
    2. Provide a clear and concise answer in the following exact format:
       - If the report is qualified:
         "Yes. The SOC 2 Type 2 Report is qualified because [specific reasons]."
       - If the report is unqualified:
         "No. The SOC 2 Type 2 Report is unqualified, indicating no reservations."
    3. Do not include any additional information or commentary beyond the specified format.
    '''

    logging.debug("Prompt for 'is_report_qualified': %s", prompt)
    response = call_ollama_api(prompt)
    return response


def qualify_soc_report(pdf_path, df_chunks_path, faiss_index_path, excel_output_path):
    """
    Performs multiple qualifier checks, then appends them to the Excel file under a "Qualifying Questions" sheet.
    Finally, appends an 'Overall SOC Viability' row at the bottom.
    """
    logging.info("Starting qualifier checks for SOC report.")

    try:
        model = SentenceTransformer('all-mpnet-base-v2')
        df_chunks = pd.read_csv(df_chunks_path)
        index = load_faiss_index(faiss_index_path)
    except Exception as e:
        logging.error("Error loading resources for qualifiers: %s", e)
        return

    # Perform the checks
    try:
        latest_report_result = is_report_latest(df_chunks, model, index)
        trust_principles_result = are_trust_principles_covered(df_chunks, model, index)
        audit_period_result = is_audit_period_sufficient(df_chunks, model, index)
        invalid_observations_result = has_invalid_observations(df_chunks, model, index)
        report_qualified_result = is_report_qualified(df_chunks, model, index)  # Updated function name
    except Exception as e:
        logging.error("Error during qualifier checks: %s", e)
        return

    qualifier_results = [
        {
            "Question": "Is the SOC 2 Type 2 Report latest (within the last 12 months)?",
            "Answer": latest_report_result
        },
        {
            "Question": "Are all three Trust Principles (Security, Availability, Confidentiality) covered?",
            "Answer": trust_principles_result
        },
        {
            "Question": "Does the SOC 2 Type 2 Report cover an audit period of at least 9 months?",
            "Answer": audit_period_result
        },
        {
            "Question": "Are there any observations in the independent auditor’s opinion that signify the report is invalid?",
            "Answer": invalid_observations_result
        },
        {
            "Question": "Is the SOC 2 Type 2 Report a qualified report?",  # Updated question text
            "Answer": report_qualified_result  # Updated variable name
        }
    ]

    def determine_status(question, answer):
        """
        Determines the status based on the question and the answer.
        For some questions, "Yes" is a Pass; for others, "Yes" is a Fail.
        """
        if "signify the report is invalid" in question or "qualified report" in question:
            # For these questions, "Yes" indicates a negative outcome (Fail)
            if answer.strip().lower().startswith("yes."):
                return "Fail"
            return "Pass"
        else:
            # For other questions, "Yes" indicates a positive outcome (Pass)
            if answer.strip().lower().startswith("yes."):
                return "Pass"
            return "Fail"

    qualifier_df = pd.DataFrame(qualifier_results)
    qualifier_df["Status"] = qualifier_df.apply(lambda row: determine_status(row["Question"], row["Answer"]), axis=1)
    qualifier_df = qualifier_df[["Question", "Status", "Answer"]]

    # Save results to the Excel
    try:
        if os.path.exists(excel_output_path):
            wb = load_workbook(excel_output_path)
            if "Qualifying Questions" in wb.sheetnames:
                ws = wb["Qualifying Questions"]
                for _, row in qualifier_df.iterrows():
                    ws.append(row.tolist())
                wb.save(excel_output_path)
            else:
                with pd.ExcelWriter(excel_output_path, engine='openpyxl', mode='a') as writer:
                    qualifier_df.to_excel(writer, sheet_name="Qualifying Questions", index=False)
        else:
            with pd.ExcelWriter(excel_output_path, mode='w', engine='openpyxl') as writer:
                qualifier_df.to_excel(writer, sheet_name="Qualifying Questions", index=False)
        logging.info("Qualifier checks appended to Excel.")
    except Exception as e:
        logging.error("Error saving qualifiers to Excel: %s", e)
        return

    # Format in Excel
    try:
        wb = load_workbook(excel_output_path)
        ws = wb["Qualifying Questions"]

        data_start_row = 2
        status_fill_pass = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        status_fill_fail = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold

        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = Font(bold=True)

        for row in range(data_start_row, ws.max_row + 1):
            status_cell = ws.cell(row=row, column=2)
            if status_cell.value == "Pass":
                status_cell.fill = status_fill_pass
            elif status_cell.value == "Fail":
                status_cell.fill = status_fill_fail

        # Overall viability row
        statuses = [ws.cell(row=r, column=2).value for r in range(data_start_row, ws.max_row + 1)]
        overall_viability = "Pass" if all(s == "Pass" for s in statuses) else "Fail"

        summary_question = "Overall SOC Viability"
        summary_status = overall_viability
        summary_answer = "SOC is valid." if overall_viability == "Pass" else "SOC is not valid."

        ws.append([summary_question, summary_status, summary_answer])
        summary_row = ws.max_row

        summary_fill = status_fill_pass if summary_status == "Pass" else status_fill_fail
        ws.cell(row=summary_row, column=2).fill = summary_fill

        bold_font = Font(bold=True)
        for col in range(1, 4):
            ws.cell(row=summary_row, column=col).font = bold_font

        wb.save(excel_output_path)
        logging.info("Excel formatting and summary row added.")
    except Exception as e:
        logging.error("Error formatting Excel for qualifiers: %s", e)
        return


if __name__ == "__main__":
    # Example usage (debug/test):
    pdf_path = "path/to/soc2_report.pdf"
    df_chunks_path = "path/to/df_qualifier_chunks.csv"
    faiss_index_path = "path/to/faiss_index_qualifiers.idx"
    excel_output_path = "path/to/output.xlsx"

    qualify_soc_report(pdf_path, df_chunks_path, faiss_index_path, excel_output_path)
