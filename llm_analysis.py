# llm_analysis.py

import os
import re
import uuid
import json
import time
import logging
import requests
import pandas as pd
from functools import wraps
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import column_index_from_string

# Configure logging with enhanced setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('llm_analysis.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def retry(exceptions, tries=3, delay=2, backoff=2):
    """
    Decorator for retrying a function call with exponential backoff.
    """
    def decorator_retry(func):
        @wraps(func)
        def wrapper_retry(*args, **kwargs):
            _tries, _delay = tries, delay
            while _tries > 1:
                try:
                    return func(*args, **kwargs)
                except exceptions as e:
                    logger.warning(f"{e}, Retrying in {_delay} seconds...")
                    time.sleep(_delay)
                    _tries -= 1
                    _delay *= backoff
            return func(*args, **kwargs)
        return wrapper_retry
    return decorator_retry

# Configuration for the Ollama API
OLLAMA_API_URL = "http://localhost:11434"   # Update if different
LLAMA_MODEL_NAME = "llama3.1"              # Update if different
MAX_TOKENS = 1024                           # Increased from 512 to 1024

@retry((requests.exceptions.RequestException, json.JSONDecodeError), tries=3, delay=2, backoff=2)
def call_ollama_api(prompt, model=LLAMA_MODEL_NAME, max_tokens=MAX_TOKENS):
    """
    Calls the Ollama API with the given prompt and returns the generated text.
    """
    logger.info("Calling Ollama API for LLM analysis.")
    url = f"{OLLAMA_API_URL}/api/generate"
    headers = {"Content-Type": "application/json"}
    session_id = str(uuid.uuid4())
    payload = {
        "model": model,
        "prompt": prompt,
        "session_id": session_id,
        "num_ctx": 2048,
        "temperature": 0.2,
        "top_p": 0.9,
        "repeat_penalty": 1.1,
        "max_tokens": max_tokens
    }

    try:
        response = requests.post(url, headers=headers, data=json.dumps(payload), stream=True, timeout=120)
        response.raise_for_status()  # Raise an exception for HTTP errors
        logger.info(f"Ollama API response status: {response.status_code}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Ollama API request failed: {e}")
        raise

    try:
        generated_text = ''
        for line in response.iter_lines():
            if line:
                line_decoded = line.decode('utf-8').strip()
                try:
                    line_json = json.loads(line_decoded)
                    token = line_json.get('response', '')
                    generated_text += token
                except json.JSONDecodeError:
                    logger.warning(f"Skipping invalid JSON line: {line_decoded}")
        logger.info("Ollama API call succeeded.")
        return generated_text.strip()
    except Exception as e:
        logger.error(f"Unexpected error during API response processing: {e}")
        raise

def load_responses(responses_path):
    """
    Loads the responses from the given Excel file.
    """
    logger.info(f"Loading responses from {responses_path}")
    try:
        df = pd.read_excel(responses_path)
        logger.info(f"Responses file '{responses_path}' loaded successfully.")
        return df
    except Exception as e:
        logger.error(f"Failed to load responses file '{responses_path}': {e}")
        raise e

def generate_analysis_prompt(control_description, answers):
    """
    Generates a prompt for the LLM to analyze control compliance based on retrieved responses,
    with strict instructions to avoid markdown, extra characters, or additional commentary.
    """
    logger.debug("Generating analysis prompt for LLM.")
    prompt = f"""You are an expert in cybersecurity compliance and controls. Your task is to analyze whether specific controls are met based on the provided responses from a document. Provide your responses in plain text only, without any markdown or formatting characters.
    
Control Description:
{control_description}

Retrieved Responses:"""
    for idx, answer in enumerate(answers, 1):
        if pd.notna(answer['text']) and answer['text']:
            prompt += f"\n{idx}. {answer['text']} (Control ID: {answer['control_id']})"

    # Strict instructions and examples
    prompt += """
    
Instructions:
1. Assess Each Response: For each retrieved response, determine the level of compliance (Fully Met, Partially Met, or Not Met).
2. Format the Output: For each response, provide an analysis in the exact format:
   1. Per review of [Retrieved Response] (Control ID: [Control_ID]), [Control Description] is Fully Met/Partially Met/Not Met. [Reasoning]

3. Final Conclusion: After analyzing all responses, provide a single-sentence conclusion starting with:
   Overall, the control "[Control Description]" is [Fully Met/Partially Met/Not Met].

Important:
- Do not include any markdown characters like asterisks (**), underscores, or backticks.
- Do not provide any additional commentary or formatting beyond the lines and single final conclusion sentence.
- If multiple responses are "Fully Met" and some are "Not Met", the overall rating may be Partially Met or Fully Met depending on the significance of the 'Not Met' gap. Use your judgment, but the final line must strictly begin with "Overall, the control ...".

Example:

Control Description:
(IDAM-01): Access to sensitive data is restricted to authorized personnel only.

Retrieved Responses:
1. The organization has implemented role-based access controls, ensuring only authorized employees can access sensitive data. (Control ID: SC-01)
2. There is no record of recent access reviews or audits. (Control ID: SC-02)

Output:
1. Per review of The organization has implemented role-based access controls, ensuring only authorized employees can access sensitive data. (Control ID: SC-01), (IDAM-01): Access to sensitive data is restricted to authorized personnel only. is Fully Met. Reasoning: Access controls are effectively implemented to restrict data access.
2. Per review of There is no record of recent access reviews or audits. (Control ID: SC-02), (IDAM-01): Access to sensitive data is restricted to authorized personnel only. is Not Met. Reasoning: Lack of recent access reviews indicates potential gaps in data security.

Follow this prompt strictly and without fail.
---
"""

    return prompt

def generate_final_conclusion_prompt(control_description, llama_analysis):
    """
    Generates a prompt for the LLM to provide a final conclusion based on the existing analysis.
    """
    prompt = f"""You are an expert in cybersecurity compliance and controls. Based on the following analysis, provide a conclusive response on whether the control is Fully Met, Partially Met, or Not Met. Respond in plain text without any markdown or formatting.

Control Description:
{control_description}

Existing Analysis:
{llama_analysis}

Instructions:
- Analyze the existing analysis based on the following rules:
  1. Fully Met: If at least one of the per-response analyses is "Fully Met", the overall conclusion should generally be "Fully Met" unless there's a critical "Not Met".
  2. Partially Met: If there are no "Fully Met" responses but one or more "Partially Met" responses, decide if they collectively fulfill the control requirements.
  3. Not Met: If no responses indicate coverage of the control or are insufficient.

- Provide a single sentence conclusion starting with "Final Conclusion:" followed by the status.
- Do not include any additional information or reasoning.

Example:
Final Conclusion: Partially Met.
"""
    return prompt

def clean_llm_output(response, control_description):
    """
    Cleans and formats the LLM output to ensure consistency.

    Parameters:
    - response (str): The raw response from the LLM.
    - control_description (str): The description of the control being analyzed.

    Returns:
    - cleaned_response (str): The formatted and cleaned response (per-response analyses).
    - individual_statuses (list): List of individual assessment statuses (Fully Met, Partially Met, Not Met).
    - final_conclusion (str): The final conclusion extracted from the LLM response.
    """
    cleaned_lines = []
    individual_statuses = []
    final_conclusion = ''

    # Split the response into lines
    lines = response.strip().split('\n')

    logger.debug(f"Raw LLM Response:\n{response}")

    for line in lines:
        # Remove leading/trailing asterisks or other unwanted characters
        line = line.strip('*').strip()
        if not line:
            continue

        # Regex to match per-response analysis (more flexible)
        per_response_match = re.match(
            r"^\d+\.\s+Per review of\s+(.+?)\s+\(Control ID:\s*([A-Za-z0-9._\- ]+)\),\s*(.+?)\s+is\s+(Fully Met|Partially Met|Not Met)\.\s*(.*)",
            line,
            re.IGNORECASE
        )
        if per_response_match:
            retrieved_response = per_response_match.group(1).strip()
            control_id = per_response_match.group(2).strip()
            control_desc = per_response_match.group(3).strip()
            status = per_response_match.group(4).strip()
            reasoning = per_response_match.group(5).strip()

            # Remove duplicate "Reasoning:" if present
            if reasoning.lower().startswith("reasoning:"):
                reasoning = reasoning[len("reasoning:"):].strip()

            # Reconstruct the line to ensure consistency without <b> tags
            formatted_line = (
                f"Per review of {control_id} {retrieved_response}; {control_desc} is {status}.\n"
                f"Reasoning: {reasoning}"
            )
            cleaned_lines.append(formatted_line)
            individual_statuses.append(status)
            continue

        # Attempt to extract the final conclusion
        final_match = re.match(
            r"(?:Overall,\s?the\s?control\s?)\"?.+?\"?\s?(?:is)\s?(Fully Met|Partially Met|Not Met)",
            line,
            re.IGNORECASE
        )
        if final_match:
            # Extract the status from final conclusion
            final_conclusion = final_match.group(1).strip()
            logger.info(f"Extracted Final Conclusion: {final_conclusion}")
            continue

        # If we see "Final Conclusion:" explicitly, try to parse status from that line
        if "Final Conclusion:" in line:
            # Attempt to match final conclusion after label
            conclusion_search = re.search(
                r"Final Conclusion:\s*(Fully Met|Partially Met|Not Met)[\.\s]*",
                line,
                re.IGNORECASE
            )
            if conclusion_search:
                final_conclusion = conclusion_search.group(1).strip()
                logger.info(f"Extracted Final Conclusion from label: {final_conclusion}")
            else:
                logger.warning(f"Could not parse final conclusion from line: {line}")
            continue

        # Unexpected format line
        logger.warning(f"Unexpected response format: {line}")

    # Join the cleaned lines with two newlines for readability
    cleaned_response = '\n\n'.join(cleaned_lines)

    # Remove duplicate Control IDs from Detailed Analysis Explanation
    cleaned_response = remove_duplicate_control_ids(cleaned_response)

    return cleaned_response, individual_statuses, final_conclusion


def remove_duplicate_control_ids(detailed_analysis):
    """
    Removes duplicate Control IDs in the same line of the Detailed Analysis Explanation.

    Parameters:
    detailed_analysis (str): The raw Detailed Analysis Explanation text.

    Returns:
    cleaned_analysis (str): The Detailed Analysis Explanation without repeated Control IDs.
    """
    lines = detailed_analysis.split('\n\n')  # Split by paragraph
    cleaned_lines = []

    for line in lines:
        # Remove repeated Control IDs in the format "AC-05 AC-05"
        cleaned_line = re.sub(r"(\b[A-Za-z0-9.-]+\b)\s+\1", r"\1", line)
        cleaned_lines.append(cleaned_line)

    cleaned_analysis = '\n\n'.join(cleaned_lines)
    return cleaned_analysis


def determine_final_conclusion(assessments):
    """
    Determines the final conclusion based on individual assessments.

    Parameters:
    - assessments (list): List of individual assessment statuses.

    Returns:
    - final_conclusion (str): The overall compliance status.
    """
    counts = {
        'Fully Met': assessments.count('Fully Met'),
        'Partially Met': assessments.count('Partially Met'),
        'Not Met': assessments.count('Not Met'),
        'Error': assessments.count('Error')
    }

    if counts['Error'] > 0:
        return 'Error in Analysis'

    # If at least one is "Fully Met", we might consider overall "Fully Met" 
    # unless there's a critical "Not Met" that can't be ignored. 
    # This can be tweaked based on your use case.
    if counts['Fully Met'] > 0 and (counts['Not Met'] == 0):
        return 'Fully Met'
    elif counts['Fully Met'] > 0 and counts['Not Met'] > 0:
        # If there's at least one "Fully Met" and one "Not Met", 
        # it might be "Partially Met" or "Fully Met" depending on severity.
        return 'Partially Met'
    elif counts['Partially Met'] > 0:
        return 'Partially Met'
    else:
        return 'Not Met'

def process_controls(df, top_k=3, max_retries=1, socketio=None, model_name=LLAMA_MODEL_NAME):
    """
    Processes each control by generating prompts, calling the LLM, cleaning the output,
    and updating the DataFrame with analysis. Implements optimized retries for incomplete responses.

    Parameters:
    - df (DataFrame): The DataFrame containing controls and retrieved answers.
    - top_k (int): Number of top responses to consider.
    - max_retries (int): Maximum number of retries for incomplete responses.
    - socketio: SocketIO instance for emitting progress updates.
    - model_name (str): The name of the model to use when calling Ollama.

    Returns:
    - df (DataFrame): Updated DataFrame with analysis.
    """
    logger.info("Processing controls with LLM analysis.")

    # Initialize columns if they don't exist
    for i in range(1, top_k + 1):
        if f'Answer_{i}' not in df.columns:
            df[f'Answer_{i}'] = ''
        if f'Answer_{i}_Control_ID' not in df.columns:
            df[f'Answer_{i}_Control_ID'] = ''

    # Initialize columns for analysis
    df['Control Status'] = ''
    df['Explanation'] = ''
    df['Final Conclusion'] = ''

    total_controls = df.shape[0]
    for idx, row in tqdm(df.iterrows(), total=total_controls, desc="Processing Controls"):
        control_description = row['Control']
        answers = []

        # Gather top_k answers
        for k in range(1, top_k + 1):
            answer_text = row.get(f'Answer_{k}', '')
            answer_control_id = row.get(f'Answer_{k}_Control_ID', '')
            if pd.notna(answer_text) and str(answer_text).strip():
                answers.append({'text': str(answer_text).strip(), 'control_id': str(answer_control_id).strip()})

        # If no answers, mark as Not Met
        if not answers:
            logger.warning(f"No valid answers found for control at index {idx}. Skipping.")
            df.at[idx, 'Control Status'] = 'Not Met'
            df.at[idx, 'Explanation'] = 'No valid responses provided.'
            df.at[idx, 'Final Conclusion'] = 'Not Met'
            if socketio:
                progress = ((idx + 1) / total_controls) * 100
                socketio.emit('progress', {'progress': progress}, broadcast=True)
            continue

        # Generate the LLM prompt
        prompt = generate_analysis_prompt(control_description, answers)

        retries = 0
        while retries <= max_retries:
            try:
                response = call_ollama_api(prompt, model=model_name, max_tokens=MAX_TOKENS)

                # Clean and format the LLM output
                cleaned_response, individual_statuses, final_conclusion_from_llm = clean_llm_output(
                    response,
                    control_description
                )

                # Determine final conclusion based on individual assessments if available
                if individual_statuses:
                    final_conclusion = determine_final_conclusion(individual_statuses)
                elif final_conclusion_from_llm:
                    # Use the final conclusion extracted from the LLM output
                    final_conclusion = final_conclusion_from_llm
                else:
                    # No statuses or final conclusion => error
                    final_conclusion = 'Error in Analysis'

                # Update the DataFrame
                df.at[idx, 'Control Status'] = final_conclusion
                df.at[idx, 'Explanation'] = cleaned_response if cleaned_response else final_conclusion_from_llm
                df.at[idx, 'Final Conclusion'] = final_conclusion

                # Emit progress update
                if socketio:
                    progress = ((idx + 1) / total_controls) * 100
                    socketio.emit('progress', {'progress': progress}, broadcast=True)

                # Break out of the retry loop after successful processing
                break

            except ValueError as ve:
                logger.error(f"Processing error at index {idx}: {ve}")
                retries += 1
                if retries > max_retries:
                    logger.error(f"Max retries exceeded for control at index {idx}. Marking as 'Error'.")
                    df.at[idx, 'Control Status'] = 'Error'
                    df.at[idx, 'Explanation'] = f"Error: {ve}"
                    df.at[idx, 'Final Conclusion'] = 'Error in Analysis'
                    if socketio:
                        progress = ((idx + 1) / total_controls) * 100
                        socketio.emit('progress', {'progress': progress}, broadcast=True)
                else:
                    logger.info(f"Retrying control at index {idx} (Attempt {retries}/{max_retries}) due to missing final conclusion...")
            except Exception as e:
                logger.error(f"Unexpected error at index {idx}: {e}")
                retries += 1
                if retries > max_retries:
                    logger.error(f"Max retries exceeded for control at index {idx}. Marking as 'Error'.")
                    df.at[idx, 'Control Status'] = 'Error'
                    df.at[idx, 'Explanation'] = f"Error: {e}"
                    df.at[idx, 'Final Conclusion'] = 'Error in Analysis'
                    if socketio:
                        progress = ((idx + 1) / total_controls) * 100
                        socketio.emit('progress', {'progress': progress}, broadcast=True)
                else:
                    logger.info(f"Retrying control at index {idx} due to unexpected error (Attempt {retries}/{max_retries})...")

    logger.info("LLM analysis completed for all controls.")
    return df

def process_final_conclusions(df, max_retries=1, socketio=None):
    """
    Processes each control's Llama Analysis to generate a final conclusion using the LLM.

    Parameters:
    - df (DataFrame): The DataFrame containing controls and their Llama Analysis.
    - max_retries (int): Maximum number of retries for incomplete responses.
    - socketio: SocketIO instance for emitting progress updates.

    Returns:
    - df (DataFrame): Updated DataFrame with final conclusions.
    """
    logger.info("Processing final conclusions with LLM.")

    if 'Final Conclusion' not in df.columns:
        df['Final Conclusion'] = ''

    total_controls = df.shape[0]
    for idx, row in tqdm(df.iterrows(), total=total_controls, desc="Finalizing Control Status"):
        control_description = row.get('User Org Control Statement', '')
        llama_analysis = row.get('Detailed Analysis Explanation', '')  # Updated to correct column

        if pd.isna(llama_analysis) or not str(llama_analysis).strip():
            logger.warning(f"No Llama Analysis found for control at index {idx}. Skipping final conclusion.")
            df.at[idx, 'Final Conclusion'] = 'Error in Analysis'
            df.at[idx, 'Control Status'] = 'Error in Analysis'
            if socketio:
                progress = ((idx + 1) / total_controls) * 100
                socketio.emit('progress', {'progress': progress}, broadcast=True)
            continue

        prompt = generate_final_conclusion_prompt(control_description, llama_analysis)
        retries = 0
        while retries <= max_retries:
            try:
                response = call_ollama_api(prompt)

                # Extract the final conclusion
                final_conclusion_match = re.match(r"Final Conclusion:\s*(Fully Met|Partially Met|Not Met)\.?$", response, re.IGNORECASE)
                if final_conclusion_match:
                    final_status = final_conclusion_match.group(1).strip()
                    df.at[idx, 'Final Conclusion'] = final_status
                    df.at[idx, 'Control Status'] = final_status
                    logger.info(f"Final Control Status for index {idx}: {final_status}")
                else:
                    logger.warning(f"Unexpected final conclusion format for control at index {idx}: {response}")
                    df.at[idx, 'Final Conclusion'] = 'Error in Analysis'
                    df.at[idx, 'Control Status'] = 'Error in Analysis'

                if socketio:
                    progress = ((idx + 1) / total_controls) * 100
                    socketio.emit('progress', {'progress': progress}, broadcast=True)

                break

            except Exception as e:
                logger.error(f"Error generating final conclusion for control at index {idx}: {e}")
                retries += 1
                if retries > max_retries:
                    logger.error(f"Max retries exceeded for final conclusion at index {idx}. Marking as 'Error in Analysis'.")
                    df.at[idx, 'Final Conclusion'] = 'Error in Analysis'
                    df.at[idx, 'Control Status'] = 'Error in Analysis'
                    if socketio:
                        progress = ((idx + 1) / total_controls) * 100
                        socketio.emit('progress', {'progress': progress}, broadcast=True)
                else:
                    logger.info(f"Retrying final conclusion for control at index {idx} (Attempt {retries}/{max_retries})...")

    logger.info("Final conclusions processed for all controls.")
    return df

def load_excel_file(file_path, sheet_name=0):
    """
    Loads the Excel file from the given path.

    Parameters:
    - file_path (str): Path to the Excel file.
    - sheet_name (str or int): Sheet name or index to load.

    Returns:
    - df (DataFrame): Loaded DataFrame.
    """
    logger.info(f"Loading Excel file from {file_path}")
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        logger.info("Framework file loaded successfully.")
        return df
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        raise e

def map_columns_by_position(framework_df):
    """
    Maps columns by their position to expected column names.

    Parameters:
    - framework_df (DataFrame): The DataFrame loaded from the framework Excel file.

    Returns:
    - framework_df (DataFrame): DataFrame with renamed columns.
    - error_message (str or None): Error message if mapping fails, else None.
    """
    logger.info("Mapping columns by position in the framework file.")
    expected_columns = ['User Org Control Domain', 'User Org Control Sub-Domain', 'User Org Control Statement']
    actual_columns = framework_df.columns.tolist()

    if len(actual_columns) < 3:
        err_msg = f"Error: Expected at least 3 columns in the framework file, found {len(actual_columns)}."
        logger.error(err_msg)
        return None, err_msg

    framework_df = framework_df.rename(columns={
        actual_columns[0]: expected_columns[0],
        actual_columns[1]: expected_columns[1],
        actual_columns[2]: expected_columns[2]
    })

    # If there are more columns than needed, slice them off
    if len(actual_columns) > 3:
        framework_df = framework_df[expected_columns]

    logger.info("Columns mapped successfully.")
    return framework_df, None

def merge_dataframes(framework_df, analysis_df):
    """
    Merges the framework DataFrame with the analysis DataFrame.

    Parameters:
    - framework_df (DataFrame): DataFrame containing framework controls.
    - analysis_df (DataFrame): DataFrame containing analysis results.

    Returns:
    - merged_df (DataFrame): Merged DataFrame.
    """
    logger.info("Merging framework DataFrame with analysis DataFrame.")
    merged_df = pd.merge(
        framework_df,
        analysis_df,
        how='left',
        left_on='User Org Control Statement',
        right_on='Control'
    )
    logger.info("DataFrames merged successfully.")
    return merged_df

def create_final_dataframe(merged_df, top_k=3):
    """
    Creates the final DataFrame with all required columns, renames columns, adds Compliance Score,
    and reorders them as specified.

    Parameters:
    - merged_df (DataFrame): Merged DataFrame containing both framework and analysis data.
    - top_k (int): Number of top answers to include.

    Returns:
    - final_df (DataFrame): Final DataFrame ready for Excel output.
    - error_message (str or None): Error message if creation fails, else None.
    """
    logger.info("Creating final DataFrame for output.")

    # Rename 'Explanation' to 'Detailed Analysis Explanation'
    if 'Explanation' in merged_df.columns:
        merged_df = merged_df.rename(columns={'Explanation': 'Detailed Analysis Explanation'})
    else:
        merged_df['Detailed Analysis Explanation'] = ""

    # Ensure 'Control Status' is not NaN
    if 'Control Status' in merged_df.columns:
        merged_df['Control Status'] = merged_df['Control Status'].fillna('Not Met')
    else:
        merged_df['Control Status'] = 'Not Met'

    # Insert 'Sr. No.' at the beginning
    merged_df = merged_df.reset_index(drop=True)
    merged_df.insert(0, 'Sr. No.', merged_df.index + 1)

    # Merge Answer_1, Answer_2, Answer_3 into 'Service Org Controls'
    def merge_controls(row):
        controls = []
        for i in range(1, top_k + 1):
            answer = row.get(f'Answer_{i}', '')
            if pd.notna(answer):
                answer_str = str(answer).strip()
                if answer_str:
                    controls.append(answer_str)
        return '\n'.join(controls)

    merged_df['Service Org Controls'] = merged_df.apply(merge_controls, axis=1)

    # Merge Answer_1_Control_ID, Answer_2_Control_ID, Answer_3_Control_ID into 'Service Org Control IDs'
    def merge_control_ids(row):
        ids = []
        for i in range(1, top_k + 1):
            cid = row.get(f'Answer_{i}_Control_ID', '')
            if pd.notna(cid):
                cid_str = str(cid).strip()
                if cid_str:
                    ids.append(cid_str)
        return '\n'.join(ids)

    merged_df['Service Org Control IDs'] = merged_df.apply(merge_control_ids, axis=1)

    # Add 'Compliance Score' column based on 'Control Status'
    status_to_score = {
        'Fully Met': 100,
        'Partially Met': 50,
        'Not Met': 0,
        'Error in Analysis': 0
    }
    merged_df['Compliance Score'] = merged_df['Control Status'].map(status_to_score).fillna(0)

    # Define the final columns in the desired order
    final_columns = [
        'Sr. No.',
        'User Org Control Domain',
        'User Org Control Sub-Domain',
        'User Org Control Statement',
        'Service Org Control IDs',
        'Service Org Controls',
        'Compliance Score',
        'Detailed Analysis Explanation',
        'Control Status'
    ]

    # Check for missing columns
    missing_final_columns = set(final_columns) - set(merged_df.columns)
    if missing_final_columns:
        err_msg = f"Error: Missing columns {missing_final_columns} after merging."
        logger.error(err_msg)
        return None, err_msg

    final_df = merged_df[final_columns]
    logger.info("Final DataFrame created successfully.")
    return final_df, None

def remove_not_met_controls(df, explanation_column="Explanation"):
    """
    Cleans "Not Met" controls in the dataframe by:
    - Removing "Not Met" entries from the specified Explanation column based on their sequence.
    - Clearing corresponding Answer_X and Answer_X_Control_ID fields.

    Args:
        df (pd.DataFrame): Input dataframe with Explanation and Answer_X columns.
        explanation_column (str): The column name containing the explanations.

    Returns:
        pd.DataFrame: Cleaned dataframe with updated fields.
    """
    for idx, row in df.iterrows():
        explanation = row.get(explanation_column, "")
        if pd.isna(explanation) or not str(explanation).strip():
            continue

        # Split the Explanation into individual control analyses based on 'Per review of'
        # The regex accounts for possible preceding newlines and ensures 'Per review of' starts each analysis
        control_analyses = re.split(r'\n*Per review of\s+', explanation)
        # The first element may be empty or partial, remove it
        control_analyses = [ca.strip() for ca in control_analyses if ca.strip()]
        retained_analyses = []

        for i, ca in enumerate(control_analyses, 1):
            # Reconstruct the 'Per review of' prefix
            ca_full = f"Per review of {ca}"
            if "is Not Met" in ca_full:
                # Clear the corresponding Answer_X and Answer_X_Control_ID fields
                answer_col = f"Answer_{i}"
                control_col = f"{answer_col}_Control_ID"
                if answer_col in df.columns and control_col in df.columns:
                    df.at[idx, answer_col] = ""
                    df.at[idx, control_col] = ""
                    logger.info(f"Cleared {answer_col} and {control_col} for control at index {idx} due to 'Not Met'.")
                else:
                    logger.warning(f"Columns {answer_col} or {control_col} not found in DataFrame for index {idx}.")
            else:
                retained_analyses.append(ca_full)

        # Update the Explanation column with retained analyses separated by two newlines
        df.at[idx, explanation_column] = "\n\n".join(retained_analyses)

    return df

def save_to_excel(df, output_path, top_k=3):
    """
    Saves the final DataFrame to an Excel file with appropriate formatting.
    Before saving, it removes any controls that were found to be 'Not Met' in the Explanation,
    and also removes those lines from the Explanation column.
    Additionally, it checks if 'Detailed Analysis Explanation' contains "Fully Met" and updates 'Control Status' accordingly.
    Then, it updates the 'Compliance Score' based on the latest 'Control Status'.
    
    Parameters:
    - df (DataFrame): The DataFrame to save.
    - output_path (str): Path where the Excel file will be saved.
    - top_k (int): Number of top answers to include.
    """
    logger.info(f"Saving final DataFrame to Excel at {output_path}")

    try:
        # First, clean "Not Met" lines
        df = remove_not_met_controls(df, explanation_column="Detailed Analysis Explanation")

        # Check if 'Detailed Analysis Explanation' column exists
        if 'Detailed Analysis Explanation' in df.columns:
            # Create a boolean mask where 'Detailed Analysis Explanation' contains "Fully Met" (case-insensitive)
            fully_met_mask = df['Detailed Analysis Explanation'].str.contains('Fully Met', case=False, na=False)
            num_fully_met = fully_met_mask.sum()

            if num_fully_met > 0:
                logger.info(f"Found 'Fully Met' in 'Detailed Analysis Explanation' for {num_fully_met} controls.")
                # Update 'Control Status' to 'Fully Met' where the mask is True
                df.loc[fully_met_mask, 'Control Status'] = 'Fully Met'
            else:
                logger.info("No instances of 'Fully Met' found in 'Detailed Analysis Explanation'.")

        else:
            logger.warning("'Detailed Analysis Explanation' column not found in DataFrame.")

        # **New Section: Update 'Compliance Score' Based on Updated 'Control Status'**
        status_to_score = {
            'Fully Met': 100,
            'Partially Met': 50,
            'Not Met': 0,
            'Error in Analysis': 0
        }
        df['Compliance Score'] = df['Control Status'].map(status_to_score).fillna(0)
        logger.info("'Compliance Score' updated based on 'Control Status'.")

        # Save the DataFrame to Excel
        df.to_excel(output_path, index=False)
        logger.info(f"DataFrame successfully saved to '{output_path}'.")

        # Load the workbook and select the active sheet
        wb = load_workbook(output_path)
        ws = wb.active

        # Define styles
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        bold_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color for headers

        # Apply formatting to header row
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill  # Apply header background color
            cell.border = border
            cell.alignment = wrap_alignment

        # Apply border and wrap text to all data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = wrap_alignment

        # Apply conditional formatting based on 'Control Status'
        control_status_col = None
        for cell in ws[1]:
            if cell.value == 'Control Status':
                control_status_col = cell.column_letter
                break

        if control_status_col:
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            ws.conditional_formatting.add(
                f'{control_status_col}2:{control_status_col}{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Not Met"'], fill=red_fill)
            )
            ws.conditional_formatting.add(
                f'{control_status_col}2:{control_status_col}{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Partially Met"'], fill=yellow_fill)
            )
            ws.conditional_formatting.add(
                f'{control_status_col}2:{control_status_col}{ws.max_row}',
                CellIsRule(operator='equal', formula=['"Fully Met"'], fill=green_fill)
            )
            logger.info("Conditional formatting applied based on 'Control Status'.")
        else:
            logger.warning("Could not find 'Control Status' column for conditional formatting.")

        # Adjust column widths based on header length
        for col in ws.columns:
            header_cell = col[0]
            if header_cell.value:
                header_length = len(str(header_cell.value))
                adjusted_width = header_length + 2
                column_letter = header_cell.column_letter
                ws.column_dimensions[column_letter].width = adjusted_width
            else:
                column_letter = col[0].column_letter
                ws.column_dimensions[column_letter].width = 15

        # Save the workbook with formatting
        wb.save(output_path)
        logger.info(f"Final Excel file saved with requested formatting to '{output_path}'.")

    except Exception as e:
        logger.error(f"Error saving the Excel file with formatting: {e}")
        raise e


# Example usage (You can remove or comment out this part if integrating into a larger system)
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="LLM Analysis Script")
    parser.add_argument('--framework', required=True, help="Path to the framework Excel file")
    parser.add_argument('--responses', required=True, help="Path to the responses Excel file")
    parser.add_argument('--output', required=True, help="Path to save the output Excel file")
    parser.add_argument('--top_k', type=int, default=3, help="Number of top responses to consider")
    args = parser.parse_args()

    try:
        framework_df = load_excel_file(args.framework)
        framework_df, error = map_columns_by_position(framework_df)
        if error:
            logger.error(error)
            exit(1)

        analysis_df = load_responses(args.responses)
        merged_df = merge_dataframes(framework_df, analysis_df)
        final_df, error = create_final_dataframe(merged_df, top_k=args.top_k)
        if error:
            logger.error(error)
            exit(1)

        processed_df = process_controls(final_df, top_k=args.top_k)
        processed_df = process_final_conclusions(processed_df)

        save_to_excel(processed_df, args.output, top_k=args.top_k)

        logger.info("LLM Analysis completed successfully.")

    except Exception as e:
        logger.error(f"An error occurred during LLM Analysis: {e}")
        exit(1)