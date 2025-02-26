import os
import logging
import threading
import time
import json
import uuid
import re
from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
from identifier import process_pdf as identify_control_ids

# Existing imports from your project
from parser import (
    generate_regex_from_sample,
    extract_text_from_pdf,
    chunk_text_by_multiple_patterns,
)
from rag import build_rag_system_with_parser, process_cybersecurity_framework_with_rag
from llm_analysis import (
    load_responses,
    process_controls,
    load_excel_file,
    map_columns_by_position,
    merge_dataframes,
    create_final_dataframe,
    save_to_excel,
    remove_not_met_controls
)
from qualifiers import (
    is_report_latest,
    are_trust_principles_covered,
    is_audit_period_sufficient,
    has_invalid_observations,
    is_report_qualified,
    describe_scope_of_services  # **New Import**
)  # Updated Import

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

import PyPDF2  # PDF processing
from sentence_transformers import SentenceTransformer
from qualifiers import qualify_soc_report
import faiss

app = Flask(__name__)
CORS(app)

# --------------------------------------------------------
# Configuration – updated paths to /Work/SOC-AI subfolders
# --------------------------------------------------------
app.config['UPLOAD_FOLDER'] = '/Work/SOC-AI/uploads'
app.config['RESULTS_FOLDER'] = '/Work/SOC-AI/results'
app.config['EXCEL_FOLDER'] = '/Work/SOC-AI/excel_outputs'
app.config['RAG_OUTPUTS'] = '/Work/SOC-AI/rag_outputs'
app.config['CHUNK_SIZE'] = 1000  # Large chunk size for better context

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['RESULTS_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXCEL_FOLDER'], exist_ok=True)
os.makedirs(app.config['RAG_OUTPUTS'], exist_ok=True)

logging.basicConfig(
    filename='LLM_analysis.log',
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

logging.info("Application started.")

# Dictionary to track progress and final result for each task_id
progress_data = {}

# -----------------------------------------------------------
# New helper functions for Section IV detection & filtering
# -----------------------------------------------------------
def is_heading(line: str, max_words: int = 12) -> bool:
    words = line.split()
    return bool(words) and len(words) <= max_words

def is_section_iv_heading(heading_text: str) -> bool:
    return heading_text.lower().strip().startswith("section iv")

def find_section_iv_page(pdf_path):
    try:
        reader = PyPDF2.PdfReader(pdf_path)
        total_pages = len(reader.pages)
        for i in range(5, total_pages):
            page_text = reader.pages[i].extract_text() or ""
            for line in page_text.splitlines():
                if is_heading(line) and is_section_iv_heading(line):
                    return i + 1
        return None
    except Exception as e:
        logging.error(f"Error detecting Section IV heading: {e}", exc_info=True)
        return None

# ----------------------------------------------------------------
# Helper functions from your existing code (unchanged)
# ----------------------------------------------------------------
def count_controls(excel_path):
    try:
        logging.info(f"Counting controls in {excel_path}.")
        if excel_path.endswith('.xlsx') or excel_path.endswith('.xls'):
            df = pd.read_excel(excel_path)
        elif excel_path.endswith('.csv'):
            df = pd.read_csv(excel_path)
        else:
            raise ValueError("Unsupported file format for counting controls.")
        num_controls = df.shape[0]
        logging.info(f"Number of controls counted: {num_controls}")
        return num_controls
    except Exception as e:
        logging.error(f"Error counting controls in {excel_path}: {e}", exc_info=True)
        return 0

def format_qualifier_sheet(excel_path):
    try:
        logging.info(f"Formatting Qualifying Questions sheet in {excel_path}.")
        wb = load_workbook(excel_path)
        if "Qualifying Questions" in wb.sheetnames:
            ws = wb["Qualifying Questions"]
            ws.column_dimensions['C'].width = 50
            data_start_row = 2
            status_fill_pass = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            status_fill_fail = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            for col in range(1, 4):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row in range(data_start_row, ws.max_row + 1):
                status_cell = ws.cell(row=row, column=2)
                if status_cell.value == "Pass":
                    status_cell.fill = status_fill_pass
                elif status_cell.value == "Fail":
                    status_cell.fill = status_fill_fail
            statuses = [ws.cell(row=r, column=2).value for r in range(data_start_row, ws.max_row + 1)]
            overall_viability = "Pass" if all(s == "Pass" for s in statuses) else "Fail"
            summary_question = "Overall SOC Viability"
            summary_status = overall_viability
            summary_answer = "SOC is valid." if overall_viability == "Pass" else "SOC is not valid."
            ws.append([summary_question, summary_status, summary_answer])
            summary_row = ws.max_row
            summary_fill = status_fill_pass if summary_status == "Pass" else status_fill_fail
            ws.cell(row=summary_row, column=2).fill = summary_fill
            bold_font = Font(bold=True)
            for col in range(1, 4):
                ws.cell(row=summary_row, column=col).font = bold_font
            ws.column_dimensions['C'].width = 50
            wb.save(excel_path)
            logging.info("Qualifying Questions sheet formatted successfully.")
        else:
            logging.warning("Qualifying Questions sheet not found in the Excel file.")
    except Exception as e:
        logging.error(f"Error formatting Qualifying Questions sheet: {e}", exc_info=True)

def chunk_text_without_patterns(text, chunk_size):
    try:
        logging.info("Starting text chunking without patterns.")
        text = text.replace('\n', ' ').replace('\r', ' ')
        sentences = re.split('(?<=[.!?]) +', text)
        chunks = []
        current_chunk = ""
        for sentence in sentences:
            current_length = len(current_chunk) + len(sentence) + 1
            if current_length <= chunk_size:
                current_chunk += (" " + sentence) if current_chunk else sentence
            else:
                chunks.append(current_chunk.strip())
                current_chunk = sentence
        if current_chunk:
            chunks.append(current_chunk.strip())
        logging.info(f"Text chunked into {len(chunks)} parts without patterns.")
        return chunks
    except Exception as e:
        logging.error(f"Error in chunk_text_without_patterns: {e}", exc_info=True)
        return []

def format_eta(seconds):
    minutes = seconds // 60
    secs = seconds % 60
    return f"{int(minutes)}m {int(secs)}s"

def sleep_seconds(task_id, seconds):
    for _ in range(seconds):
        time.sleep(1)
        if task_id in progress_data:
            if progress_data[task_id].get('cancelled'):
                logging.info(f"Task {task_id} has been cancelled during sleep.")
                break
            if progress_data[task_id]['eta'] > 0:
                progress_data[task_id]['eta'] -= 1

def rename_sheet_to_soc_mapping(excel_path):
    try:
        logging.info(f"Renaming 'Sheet1' to 'Control Assessment' in {excel_path}.")
        wb = load_workbook(excel_path)
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
            ws.title = "Control Assessment"
            wb.save(excel_path)
            logging.info(f"Renamed 'Sheet1' to 'Control Assessment' in {excel_path}.")
        else:
            logging.warning("Sheet1 not found in the Excel file.")
    except Exception as e:
        logging.error(f"Error renaming Sheet1 to Control Assessment: {e}", exc_info=True)

def format_compliance_sheet(excel_path):
    try:
        logging.info(f"Formatting Compliance Score sheet in {excel_path}.")
        wb = load_workbook(excel_path)
        if "Compliance Score" not in wb.sheetnames:
            logging.warning("Compliance Score sheet not found for formatting.")
            return
        ws = wb["Compliance Score"]
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        domain_col = None
        domain_score_col = None
        subdomain_col = None
        subdomain_score_col = None
        overall_col = None
        for i, cell in enumerate(ws[1], start=1):
            if cell.value == "Domain Name":
                domain_col = i
            elif cell.value == "Domain Compliance Score":
                domain_score_col = i
            elif cell.value == "Sub-Domain Name":
                subdomain_col = i
            elif cell.value == "Sub-Domain Compliance Score":
                subdomain_score_col = i
            elif cell.value == "Overall Compliance Score":
                overall_col = i
        if not all([domain_col, domain_score_col, subdomain_col, subdomain_score_col, overall_col]):
            logging.error("One or more required columns not found in Compliance Score sheet.")
            return
        max_row = ws.max_row
        if max_row > 1:
            ws.merge_cells(start_row=2, start_column=overall_col, end_row=max_row, end_column=overall_col)
        current_domain = None
        domain_start_row = 2
        for row in range(2, max_row + 1):
            cell_value = ws.cell(row=row, column=domain_col).value
            if cell_value != current_domain:
                if current_domain is not None:
                    ws.merge_cells(
                        start_row=domain_start_row,
                        start_column=domain_col,
                        end_row=row - 1,
                        end_column=domain_col
                    )
                    ws.merge_cells(
                        start_row=domain_start_row,
                        start_column=domain_score_col,
                        end_row=row - 1,
                        end_column=domain_score_col
                    )
                current_domain = cell_value
                domain_start_row = row
        if current_domain is not None and domain_start_row < max_row:
            ws.merge_cells(
                start_row=domain_start_row,
                start_column=domain_col,
                end_row=max_row,
                end_column=domain_col
            )
            ws.merge_cells(
                start_row=domain_start_row,
                start_column=domain_score_col,
                end_row=max_row,
                end_column=domain_score_col
            )
        ws.column_dimensions[get_column_letter(domain_col)].width = 25
        ws.column_dimensions[get_column_letter(domain_score_col)].width = 25
        ws.column_dimensions[get_column_letter(subdomain_col)].width = 25
        ws.column_dimensions[get_column_letter(subdomain_score_col)].width = 25
        ws.column_dimensions[get_column_letter(overall_col)].width = 25
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        bold_font = Font(bold=True, color="000000")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wb.save(excel_path)
        logging.info("Compliance Score sheet formatted successfully.")
    except Exception as e:
        logging.error(f"Error formatting Compliance Score sheet: {e}", exc_info=True)

def create_executive_summary(input_excel_path, output_excel_path):
    try:
        logging.info("Generating Executive Summary...")
        df = pd.read_excel(input_excel_path, sheet_name='Control Assessment')
        logging.info(f"Columns in 'Control Assessment': {list(df.columns)}")
        expected_columns = [
            'Sr. No.',
            'User Org Control Domain',
            'User Org Control Sub-Domain',
            'User Org Control Statement',
            'Service Org Control IDs',
            'Service Org Controls',
            'Compliance Score',
            'Detailed Analysis Explanation',
            'Control Status'
        ]
        for col in expected_columns:
            if col not in df.columns:
                raise ValueError(f"Column '{col}' not found in 'Control Assessment'.")
        df['User Org Control Domain'] = df['User Org Control Domain'].astype(str).str.strip().str.title()
        df['User Org Control Sub-Domain'] = df['User Org Control Sub-Domain'].astype(str).str.strip().str.title()
        unique_domains = df['User Org Control Domain'].unique()
        unique_subdomains = df['User Org Control Sub-Domain'].unique()
        domain_avg = df.groupby("User Org Control Domain")["Compliance Score"].mean().reset_index()
        domain_avg["Compliance Score"] = domain_avg["Compliance Score"] / 100.0
        overall_avg = domain_avg["Compliance Score"].mean()
        domain_status_counts = df.groupby(["User Org Control Domain", "Control Status"]).size().unstack(fill_value=0).reset_index()
        for status in ["Fully Met", "Partially Met", "Not Met"]:
            if status not in domain_status_counts.columns:
                domain_status_counts[status] = 0
        domain_summary = domain_avg.merge(domain_status_counts, on="User Org Control Domain", how="left")
        if len(domain_summary) != len(unique_domains):
            missing_domains = set(unique_domains) - set(domain_summary['User Org Control Domain'])
            for domain in missing_domains:
                domain_summary = domain_summary.append({
                    "User Org Control Domain": domain,
                    "Compliance Score": 0.0,
                    "Fully Met": 0,
                    "Partially Met": 0,
                    "Not Met": 0
                }, ignore_index=True)
        subdomain_avg = df.groupby("User Org Control Sub-Domain")["Compliance Score"].mean().reset_index()
        subdomain_avg["Compliance Score"] = subdomain_avg["Compliance Score"] / 100.0
        subdomain_status_counts = df.groupby(["User Org Control Sub-Domain", "Control Status"]).size().unstack(fill_value=0).reset_index()
        for status in ["Fully Met", "Partially Met", "Not Met"]:
            if status not in subdomain_status_counts.columns:
                subdomain_status_counts[status] = 0
        subdomain_summary = subdomain_avg.merge(subdomain_status_counts, on="User Org Control Sub-Domain", how="left")
        if len(subdomain_summary) != len(unique_subdomains):
            missing_subdomains = set(unique_subdomains) - set(subdomain_summary['User Org Control Sub-Domain'])
            for subd in missing_subdomains:
                subdomain_summary = subdomain_summary.append({
                    "User Org Control Sub-Domain": subd,
                    "Compliance Score": 0.0,
                    "Fully Met": 0,
                    "Partially Met": 0,
                    "Not Met": 0
                }, ignore_index=True)
        try:
            qualifying_df = pd.read_excel(input_excel_path, sheet_name='Qualifying Questions')
            logging.info(f"Columns in 'Qualifying Questions': {list(qualifying_df.columns)}")
            status_column = qualifying_df.columns[1]
            fail_exists = qualifying_df[status_column].astype(str).str.strip().str.lower().eq('fail').any()
            soc_usability_status = "No" if fail_exists else "Yes"
            logging.info(f"SOC Usability Status determined as: {soc_usability_status}")
        except Exception as e:
            logging.error(f"Error reading 'Qualifying Questions' sheet: {e}", exc_info=True)
            soc_usability_status = "No"
        wb = load_workbook(input_excel_path)
        if "Executive Summary" in wb.sheetnames:
            del wb["Executive Summary"]
        ws = wb.create_sheet("Executive Summary")
        wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index("Executive Summary")))
        legend_colors = {
            ">90%": "CCFFCC",
            "<90% & >75%": "FFF2CC",
            "<75%": "FFCCCC"
        }
        fill_legend = {
            ">90%": PatternFill(start_color=legend_colors[">90%"], end_color=legend_colors[">90%"], fill_type="solid"),
            "<90% & >75%": PatternFill(start_color=legend_colors["<90% & >75%"], end_color=legend_colors["<90% & >75%"], fill_type="solid"),
            "<75%": PatternFill(start_color=legend_colors["<75%"], end_color=legend_colors["<75%"], fill_type="solid")
        }
        compliance_fill_colors = {
            "high": legend_colors[">90%"],
            "medium": legend_colors["<90% & >75%"],
            "low": legend_colors["<75%"]
        }
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        thin_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells('B2:C2')
        ws['B2'] = "Legend"
        ws['B2'].font = Font(bold=True, size=14, color="000000")
        ws['B2'].alignment = center_alignment
        ws['B2'].fill = header_fill
        legend_items = [
            { "text": ">90%", "fill": fill_legend[">90%"] },
            { "text": "<90% & >75%", "fill": fill_legend["<90% & >75%"] },
            { "text": "<75%", "fill": fill_legend["<75%"] }
        ]
        for idx, item in enumerate(legend_items, start=3):
            ws[f'B{idx}'] = item["text"]
            ws[f'B{idx}'].alignment = left_alignment
            ws[f'B{idx}'].font = Font(bold=False, color="000000")
            ws[f'C{idx}'] = ""
            ws[f'C{idx}'].fill = item["fill"]
            ws[f'C{idx}'].border = thin_border
        for row in range(2, 6):
            for col in ['B', 'C']:
                ws[f'{col}{row}'].border = thin_border
        ws['B7'] = "SOC Usability"
        ws['B7'].font = Font(bold=True, color="000000")
        ws['B7'].alignment = left_alignment
        ws['B7'].fill = header_fill
        ws['B7'].border = thin_border
        ws['C7'] = soc_usability_status
        ws['C7'].number_format = "@"
        ws['C7'].alignment = center_alignment
        ws['C7'].border = thin_border
        soc_usability_colors = {
            "yes": "CCFFCC",
            "no": "FFCCCC"
        }
        soc_usability_value = ws['C7'].value
        if soc_usability_value:
            color_key = soc_usability_value.strip().lower()
            fill_color_soc = soc_usability_colors.get(color_key)
            if fill_color_soc:
                ws['C7'].fill = PatternFill(start_color=fill_color_soc, end_color=fill_color_soc, fill_type="solid")
        ws['B10'] = "Overall Compliance Percentage"
        ws['B10'].font = Font(bold=True, color="000000")
        ws['B10'].alignment = left_alignment
        ws['B10'].fill = header_fill
        ws['B10'].border = thin_border
        ws['C10'] = overall_avg
        ws['C10'].number_format = "0.00%"
        ws['C10'].alignment = center_alignment
        ws['C10'].border = thin_border
        compliance_percentage_overall = overall_avg * 100
        if compliance_percentage_overall >= 90:
            fill_color_overall = compliance_fill_colors["high"]
        elif 75 <= compliance_percentage_overall < 90:
            fill_color_overall = compliance_fill_colors["medium"]
        else:
            fill_color_overall = compliance_fill_colors["low"]
        ws['C10'].fill = PatternFill(start_color=fill_color_overall, end_color=fill_color_overall, fill_type="solid")
        ws['B12'] = "Domain-wise Compliance Breakdown"
        ws['B12'].font = Font(bold=True, size=12, color="000000")
        ws['B12'].alignment = left_alignment
        ws['B12'].fill = header_fill
        ws['B12'].border = thin_border
        headers_domain = ["Domain", "Average Compliance Score", "Fully Met Controls", "Partially Met Controls", "Not Met Controls"]
        start_col = 2
        for idx, header in enumerate(headers_domain, start=start_col):
            cell = ws.cell(row=12, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        start_row_domain = 13
        for i, row_data in domain_summary.iterrows():
            row_num = start_row_domain + i
            ws.cell(row=row_num, column=2, value=row_data["User Org Control Domain"]).alignment = left_alignment
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=3, value=row_data["Compliance Score"])
            ws.cell(row=row_num, column=3).number_format = "0.00%"
            ws.cell(row=row_num, column=3).alignment = center_alignment
            ws.cell(row=row_num, column=3).border = thin_border
            ws.cell(row=row_num, column=4, value=row_data.get("Fully Met", 0)).alignment = center_alignment
            ws.cell(row=row_num, column=4).border = thin_border
            ws.cell(row=row_num, column=5, value=row_data.get("Partially Met", 0)).alignment = center_alignment
            ws.cell(row=row_num, column=5).border = thin_border
            ws.cell(row=row_num, column=6, value=row_data.get("Not Met", 0)).alignment = center_alignment
            ws.cell(row=row_num, column=6).border = thin_border
            c_perc = row_data["Compliance Score"] * 100
            if c_perc >= 90:
                fill_color = compliance_fill_colors["high"]
            elif 75 <= c_perc < 90:
                fill_color = compliance_fill_colors["medium"]
            else:
                fill_color = compliance_fill_colors["low"]
            ws.cell(row=row_num, column=3).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        last_domain_row = start_row_domain + len(domain_summary) - 1
        start_row_subdomain_header = last_domain_row + 2
        ws.cell(row=start_row_subdomain_header, column=2, value="Sub-Domain-wise Compliance Breakdown")
        ws.cell(row=start_row_subdomain_header, column=2).font = Font(bold=True, size=12, color="000000")
        ws.cell(row=start_row_subdomain_header, column=2).alignment = left_alignment
        ws.cell(row=start_row_subdomain_header, column=2).fill = header_fill
        ws.cell(row=start_row_subdomain_header, column=2).border = thin_border
        headers_subdomain = ["Sub-Domain", "Average Compliance Score", "Fully Met Controls", "Partially Met Controls", "Not Met Controls"]
        start_col_sub = 2
        for idx, header in enumerate(headers_subdomain, start=start_col_sub):
            cell = ws.cell(row=start_row_subdomain_header, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        start_row_subdomain = start_row_subdomain_header + 1
        for i, row_data in subdomain_summary.iterrows():
            row_num = start_row_subdomain + i
            ws.cell(row=row_num, column=2, value=row_data["User Org Control Sub-Domain"]).alignment = left_alignment
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=3, value=row_data["Compliance Score"])
            ws.cell(row=row_num, column=3).number_format = "0.00%"
            ws.cell(row=row_num, column=3).alignment = center_alignment
            ws.cell(row=row_num, column=3).border = thin_border
            ws.cell(row=row_num, column=4, value=row_data.get("Fully Met", 0)).alignment = center_alignment
            ws.cell(row=row_num, column=4).border = thin_border
            ws.cell(row=row_num, column=5, value=row_data.get("Partially Met", 0)).alignment = center_alignment
            ws.cell(row=row_num, column=5).border = thin_border
            ws.cell(row=row_num, column=6, value=row_data.get("Not Met", 0)).alignment = center_alignment
            ws.cell(row=row_num, column=6).border = thin_border
            c_perc = row_data["Compliance Score"] * 100
            if c_perc >= 90:
                fill_color = compliance_fill_colors["high"]
            elif 75 <= c_perc < 90:
                fill_color = compliance_fill_colors["medium"]
            else:
                fill_color = compliance_fill_colors["low"]
            ws.cell(row=row_num, column=3).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        def autofit_columns(ws, min_width=10, max_width=50):
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                adjusted_width = length + 2
                adjusted_width = max(min_width, min(adjusted_width, max_width))
                column_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[column_letter].width = adjusted_width
        autofit_columns(ws)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        wb.save(output_excel_path)
        wb.close()
        logging.info(f"Executive Summary created and saved to {output_excel_path}.")
    except Exception as e:
        logging.error(f"Error in create_executive_summary: {e}", exc_info=True)
        raise

def create_cuec_sheet(pdf_path, summary_output_path, pages_to_skip=5):
    """
    Processes the complementary content via the CUEC module and adds a new sheet
    named 'Complementary User Entity Controls' to the final Excel file.
    """
    try:
        from pathlib import Path
        import CUEC  # Ensure CUEC.py is available in the PYTHONPATH
        df_cuec = CUEC.process_pdf_to_dataframe(Path(pdf_path), pages_to_skip=pages_to_skip)
        wb = load_workbook(summary_output_path)
        sheet_names = wb.sheetnames
        try:
            control_assessment_index = sheet_names.index("Control Assessment")
        except ValueError:
            control_assessment_index = 0
        ws_cuec = wb.create_sheet("CUECs")
        sheets = wb._sheets
        sheets.remove(ws_cuec)
        sheets.insert(control_assessment_index + 1, ws_cuec)
        from openpyxl.utils.dataframe import dataframe_to_rows
        header = list(df_cuec.columns)
        ws_cuec.append(header)
        for row in dataframe_to_rows(df_cuec, index=False, header=False):
            ws_cuec.append(row)
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        thin_border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws_cuec[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        for col in ws_cuec.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = max(10, max_length + 2)
            ws_cuec.column_dimensions[column].width = adjusted_width
        wb.save(summary_output_path)
        wb.close()
        logging.info(f"CUECs sheet added to {summary_output_path}")
    except Exception as e:
        logging.error(f"Error in create_cuec_sheet: {e}", exc_info=True)
        raise

def detect_control_id_pages(pdf_path, regex_to_cids):
    control_id_pages = {cid: [] for cids in regex_to_cids.values() for cid in cids}
    try:
        logging.info(f"Detecting pages for Control IDs in {pdf_path}.")
        reader = PyPDF2.PdfReader(pdf_path)
        for page_num, page in enumerate(reader.pages, start=1):
            text = page.extract_text()
            if text:
                for regex, cids in regex_to_cids.items():
                    if re.search(regex, text, re.IGNORECASE):
                        for cid in cids:
                            if page_num not in control_id_pages[cid]:
                                control_id_pages[cid].append(page_num)
                                logging.debug(f"Found Control ID '{cid}' on page {page_num}.")
        logging.info(f"Control ID pages detected: {control_id_pages}")
        return control_id_pages
    except Exception as e:
        logging.error(f"Error in detect_control_id_pages: {e}", exc_info=True)
        return control_id_pages

def determine_page_range(control_id_pages, regex_to_cids):
    start_pages = []
    for regex, cids in regex_to_cids.items():
        pages = [min(control_id_pages[cid]) for cid in cids if control_id_pages.get(cid)]
        if pages:
            earliest_page = min(pages)
            start_pages.append(earliest_page)
    start_page = min(start_pages) if start_pages else 1
    all_pages = [p for pages in control_id_pages.values() for p in pages]
    end_page = max(all_pages) if all_pages else 1
    return (start_page, end_page)

# -------------------------------------------------------------------
# Background processing function for /process_all endpoint
# -------------------------------------------------------------------
def background_process(task_id, pdf_path, excel_path, start_page, end_page, control_id,
                       soc_report_filename, framework_filename):
    try:
        logging.info(f"Background processing started for task_id: {task_id}.")
        if start_page < 1 or end_page < start_page:
            raise ValueError("Invalid page range: ensure that 1 <= start_page <= end_page.")
        qualifiers_start_page = 1
        qualifiers_end_page = start_page - 1 if start_page > 1 else 1
        pre_llm_time = 40
        pre_llm_steps = 6
        pre_llm_step_time = int(pre_llm_time / pre_llm_steps)
        qualifier_time = 21
        llm_time_per_control = 7
        num_controls = count_controls(excel_path)
        llm_time = num_controls * llm_time_per_control
        total_eta = pre_llm_time + llm_time + qualifier_time
        progress_data[task_id]['eta'] = int(total_eta)
        progress_data[task_id]['progress'] = 0.0
        progress_data[task_id]['status'] = "Initializing..."
        progress_data[task_id]['download_url'] = None
        progress_data[task_id]['error'] = None
        progress_data[task_id]['num_controls'] = num_controls
        progress_data[task_id]['cancelled'] = False
        from threading import Lock
        progress_lock = Lock()
        def check_cancel(tid):
            if progress_data[tid].get('cancelled'):
                logging.info(f"Task {tid} has been cancelled.")
                with progress_lock:
                    progress_data[tid]['status'] = "Cancelled"
                    progress_data[tid]['progress'] = 0
                    progress_data[tid]['eta'] = 0
                    progress_data[tid]['error'] = "Task was cancelled by the user."
                raise Exception("Task cancelled by user.")
        progress_increment_pre_llm = 20.0 / pre_llm_steps
        progress_increment_llm = 70.0 / num_controls if num_controls > 0 else 0
        progress_increment_qualifier = 10.0
        pre_llm_phase_descriptions = [
            "Extracting full text for qualifiers...",
            "Chunking full text for qualifiers...",
            "Building RAG system for qualifiers...",
            "Extracting controls' text...",
            "Chunking controls' text...",
            "Building RAG system for controls..."
        ]
        if not control_id.strip():
            raise ValueError("No Control IDs were selected. Please provide Control IDs before continuing.")
        for idx, step_description in enumerate(pre_llm_phase_descriptions):
            check_cancel(task_id)
            with progress_lock:
                progress_data[task_id]['status'] = step_description
            logging.info(f"Task {task_id}: {step_description}, ETA: {format_eta(progress_data[task_id]['eta'])}")
            if idx == 0:
                full_text_output_path = os.path.join(app.config['RAG_OUTPUTS'], f"full_text_{task_id}.txt")
                full_extracted_text = extract_text_from_pdf(pdf_path, qualifiers_start_page, qualifiers_end_page, full_text_output_path)
                sleep_seconds(task_id, pre_llm_step_time)
                if not os.path.exists(full_text_output_path):
                    raise FileNotFoundError(f"Extracted full text file not found at {full_text_output_path}")
                with progress_lock:
                    progress_data[task_id]['progress'] += progress_increment_pre_llm
            elif idx == 1:
                chunk_size = app.config['CHUNK_SIZE']
                text_chunks = chunk_text_without_patterns(full_extracted_text, chunk_size)
                df_qualifier_chunks_path = os.path.join(app.config['RAG_OUTPUTS'], f"df_qualifier_chunks_{task_id}.csv")
                pd.DataFrame({"Content": text_chunks}).to_csv(df_qualifier_chunks_path, index=False)
                sleep_seconds(task_id, pre_llm_step_time)
                if not os.path.exists(df_qualifier_chunks_path):
                    raise FileNotFoundError(f"Qualifier chunks file not found at {df_qualifier_chunks_path}")
                with progress_lock:
                    progress_data[task_id]['progress'] += progress_increment_pre_llm
            elif idx == 2:
                faiss_index_qualifiers_path = os.path.join(app.config['RAG_OUTPUTS'], f"faiss_index_qualifiers_{task_id}.idx")
                build_rag_system_with_parser(
                    pdf_path=pdf_path,
                    start_page=qualifiers_start_page,
                    end_page=qualifiers_end_page,
                    control_patterns=None,
                    output_text_path=full_text_output_path,
                    df_chunks_path=df_qualifier_chunks_path,
                    faiss_index_path=faiss_index_qualifiers_path,
                    chunk_size=chunk_size
                )
                sleep_seconds(task_id, pre_llm_step_time)
                if not os.path.exists(faiss_index_qualifiers_path):
                    raise FileNotFoundError(f"Qualifiers FAISS index not found at {faiss_index_qualifiers_path}")
                with progress_lock:
                    progress_data[task_id]['progress'] += progress_increment_pre_llm
            elif idx == 3:
                controls_output_text_path = os.path.join(app.config['RAG_OUTPUTS'], f"controls_text_{task_id}.txt")
                controls_extracted_text = extract_text_from_pdf(pdf_path, start_page, end_page, controls_output_text_path)
                sleep_seconds(task_id, pre_llm_step_time)
                if not os.path.exists(controls_output_text_path):
                    raise FileNotFoundError(f"Controls extracted text file not found at {controls_output_text_path}")
                with progress_lock:
                    progress_data[task_id]['progress'] += progress_increment_pre_llm
            elif idx == 4:
                control_ids_raw = control_id
                control_ids = [cid.strip() for cid in control_ids_raw.split(',') if cid.strip()]
                control_patterns = [generate_regex_from_sample(cid) for cid in control_ids]
                logging.info(f"Generated regex patterns: {control_patterns}")
                control_text_chunks = chunk_text_by_multiple_patterns(controls_extracted_text, control_patterns)
                df_control_chunks_path = os.path.join(app.config['RAG_OUTPUTS'], f"df_control_chunks_{task_id}.csv")
                control_text_chunks.to_csv(df_control_chunks_path, index=False)
                sleep_seconds(task_id, pre_llm_step_time)
                if not os.path.exists(df_control_chunks_path):
                    raise FileNotFoundError(f"Control chunks file not found at {df_control_chunks_path}")
                with progress_lock:
                    progress_data[task_id]['progress'] += progress_increment_pre_llm
            elif idx == 5:
                faiss_index_controls_path = os.path.join(app.config['RAG_OUTPUTS'], f"faiss_index_controls_{task_id}.idx")
                build_rag_system_with_parser(
                    pdf_path=pdf_path,
                    start_page=start_page,
                    end_page=end_page,
                    control_patterns=control_patterns,
                    output_text_path=controls_output_text_path,
                    df_chunks_path=df_control_chunks_path,
                    faiss_index_path=faiss_index_controls_path,
                    chunk_size=app.config['CHUNK_SIZE']
                )
                sleep_seconds(task_id, pre_llm_step_time)
                if not os.path.exists(faiss_index_controls_path):
                    raise FileNotFoundError(f"Controls FAISS index not found at {faiss_index_controls_path}")
                with progress_lock:
                    progress_data[task_id]['progress'] += progress_increment_pre_llm
        check_cancel(task_id)
        with progress_lock:
            progress_data[task_id]['status'] = "Processing control framework with RAG..."
        logging.info(f"Task {task_id}: Processing control framework with RAG. ETA: {format_eta(progress_data[task_id]['eta'])}")
        processed_framework_path = os.path.join(app.config['RAG_OUTPUTS'], f"cybersecurity_framework_with_answers_{task_id}.xlsx")
        faiss_index_controls_path = os.path.join(app.config['RAG_OUTPUTS'], f"faiss_index_controls_{task_id}.idx")
        df_control_chunks_path = os.path.join(app.config['RAG_OUTPUTS'], f"df_control_chunks_{task_id}.csv")
        process_cybersecurity_framework_with_rag(
            excel_input_path=excel_path,
            output_path=processed_framework_path,
            faiss_index_path=faiss_index_controls_path,
            df_chunks_path=df_control_chunks_path,
            top_k=3
        )
        check_cancel(task_id)
        if not os.path.exists(processed_framework_path):
            raise FileNotFoundError(f"Processed framework file not found at {processed_framework_path}")
        with progress_lock:
            progress_data[task_id]['status'] = "Analyzing controls with LLM..."
        logging.info(f"Task {task_id}: Analyzing controls with LLM. ETA: {format_eta(progress_data[task_id]['eta'])}")
        analysis_df = load_responses(processed_framework_path)
        analyzed_rows = []
        for i, row in analysis_df.iterrows():
            check_cancel(task_id)
            with progress_lock:
                progress_data[task_id]['status'] = f"Analyzing control {i+1} of {num_controls} with LLM..."
            logging.info(f"Task {task_id}: Analyzing control {i+1}/{num_controls}. ETA: {format_eta(progress_data[task_id]['eta'])}")
            sleep_seconds(task_id, llm_time_per_control)
            processed_row = process_controls(pd.DataFrame([row]))
            analyzed_rows.append(processed_row)
            with progress_lock:
                progress_data[task_id]['progress'] += progress_increment_llm
        analyzed_df = pd.concat(analyzed_rows, ignore_index=True) if analyzed_rows else pd.DataFrame()
        analyzed_df = remove_not_met_controls(analyzed_df)
        analysis_output_path = os.path.join(app.config['RESULTS_FOLDER'], f"Framework_Analysis_Completed_{task_id}.xlsx")
        analyzed_df.to_excel(analysis_output_path, index=False)
        logging.info(f"Task {task_id}: LLM analysis completed at {analysis_output_path}")
        check_cancel(task_id)
        framework_df = load_excel_file(excel_path)
        framework_df, error = map_columns_by_position(framework_df)
        if error:
            raise ValueError(error)
        merged_df = merge_dataframes(framework_df, analyzed_df)
        final_df, error = create_final_dataframe(merged_df)
        if error:
            raise ValueError(error)
        final_output_path = os.path.join(app.config['EXCEL_FOLDER'], f"Final_Control_Status_{task_id}.xlsx")
        save_to_excel(final_df, final_output_path)
        logging.info(f"Task {task_id}: Final control status saved to {final_output_path}")
        with progress_lock:
            progress_data[task_id]['progress'] = 90.0
        rename_sheet_to_soc_mapping(final_output_path)
        check_cancel(task_id)
        with progress_lock:
            progress_data[task_id]['status'] = "Performing qualifier checks..."
        logging.info(f"Task {task_id}: Performing qualifier checks. ETA: {format_eta(progress_data[task_id]['eta'])}")
        try:
            qualify_soc_report(
                pdf_path=pdf_path,
                df_chunks_path=os.path.join(app.config['RAG_OUTPUTS'], f"df_qualifier_chunks_{task_id}.csv"),
                faiss_index_path=os.path.join(app.config['RAG_OUTPUTS'], f"faiss_index_qualifiers_{task_id}.idx"),
                excel_output_path=final_output_path
            )
            logging.info(f"Task {task_id}: Qualifier checks completed.")
        except Exception as q_ex:
            logging.error(f"Error during qualifier checks: {q_ex}", exc_info=True)
        format_qualifier_sheet(final_output_path)
        with progress_lock:
            progress_data[task_id]['status'] = "Finalizing qualifier processing..."
        logging.info(f"Task {task_id}: Finalizing qualifier processing. ETA: {format_eta(progress_data[task_id]['eta'])}")
        sleep_seconds(task_id, qualifier_time)
        with progress_lock:
            progress_data[task_id]['eta'] -= qualifier_time
            progress_data[task_id]['progress'] += progress_increment_qualifier
        check_cancel(task_id)
        with progress_lock:
            progress_data[task_id]['status'] = "Creating Executive Summary..."
        logging.info(f"Task {task_id}: Creating Executive Summary. ETA: {format_eta(progress_data[task_id]['eta'])}")
        soc_report_basename = os.path.splitext(os.path.basename(soc_report_filename))[0]
        framework_basename = os.path.splitext(os.path.basename(framework_filename))[0]
        final_filename = f"{soc_report_basename}_Baselined_Vs_{framework_basename}.xlsx"
        summary_output_path = os.path.join(app.config['EXCEL_FOLDER'], final_filename)
        create_executive_summary(final_output_path, summary_output_path)
        logging.info(f"Task {task_id}: Executive Summary created at {summary_output_path}")
        # ----------------------------------------------------------------
        # NEW STEP: Create Complementary User Entity Controls sheet via CUEC
        # ----------------------------------------------------------------
        with progress_lock:
            progress_data[task_id]['status'] = "Creating Complementary User Entity Controls sheet..."
        logging.info(f"Task {task_id}: Creating Complementary User Entity Controls sheet.")
        create_cuec_sheet(pdf_path, summary_output_path, pages_to_skip=5)
        # ----------------------------------------------------------------
        with progress_lock:
            progress_data[task_id]['download_url'] = f"https://91upn2obiudwqc-5000.proxy.runpod.net/download/{final_filename}"
            progress_data[task_id]['progress'] = 100.0
            progress_data[task_id]['eta'] = 0
            progress_data[task_id]['status'] = "Task completed successfully."
        logging.info(f"Task {task_id}: Task completed successfully.")
    except Exception as e:
        logging.error(f"Error in background_process (task_id: {task_id}): {e}", exc_info=True)
        from threading import Lock
        progress_lock = Lock()
        if "cancelled" in str(e).lower():
            with progress_lock:
                progress_data[task_id]['cancelled'] = True
                progress_data[task_id]['status'] = "Cancelled"
                progress_data[task_id]['progress'] = 0
                progress_data[task_id]['eta'] = 0
                progress_data[task_id]['error'] = "Task was cancelled by the user."
        else:
            with progress_lock:
                progress_data[task_id]['status'] = "Error occurred."
                progress_data[task_id]['error'] = str(e)
                progress_data[task_id]['eta'] = 0

def determine_status(question, answer):
    if "signify the report is invalid" in question or "qualified report" in question:
        if answer.strip().lower().startswith("yes."):
            return "Fail"
        return "Pass"
    else:
        if answer.strip().lower().startswith("yes."):
            return "Pass"
        return "Fail"

# -------------------------------------------------------
# UPDATED Endpoint: /initial_qualifier_check (unchanged)
# -------------------------------------------------------
@app.route('/initial_qualifier_check', methods=['POST'])
def initial_qualifier_check():
    logging.info("Received request to /initial_qualifier_check endpoint.")
    try:
        pdf_file = request.files.get('pdf_file')
        if not pdf_file:
            raise ValueError("Missing required file: pdf_file")
        filename_pdf = secure_filename(pdf_file.filename)
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename_pdf)
        pdf_file.save(pdf_path)
        logging.info(f"PDF file saved to {pdf_path}")
        reader = PyPDF2.PdfReader(pdf_path)
        total_pages = len(reader.pages)
        full_text = ""
        for page_num in range(total_pages):
            page_text = reader.pages[page_num].extract_text()
            if page_text:
                full_text += "\n" + page_text
        chunk_size = app.config['CHUNK_SIZE']
        text_chunks = chunk_text_without_patterns(full_text, chunk_size)
        df_temp = pd.DataFrame({"Content": text_chunks})
        model = SentenceTransformer('all-mpnet-base-v2')
        embeddings = model.encode(df_temp["Content"].tolist(), show_progress_bar=False).astype('float32')
        dimension = embeddings.shape[1]
        index = faiss.IndexFlatIP(dimension)
        index.add(embeddings)
        latest_report_result = is_report_latest(df_temp, model, index, top_k=3)
        trust_principles_result = are_trust_principles_covered(df_temp, model, index, top_k=3)
        audit_period_result = is_audit_period_sufficient(df_temp, model, index, top_k=3)
        invalid_observations_result = has_invalid_observations(df_temp, model, index, top_k=3)
        report_qualified_result = is_report_qualified(df_temp, model, index, top_k=3)
        scope_of_services_result = describe_scope_of_services(df_temp, model, index, top_k=3)
        qualifiers = [
            {
                "question": "Is the SOC 2 Type 2 Report latest (within the last 12 months)?",
                "answer": latest_report_result,
                "status": determine_status("Is the SOC 2 Type 2 Report latest?", latest_report_result)
            },
            {
                "question": "Are all three Trust Principles (Security, Availability, Confidentiality) covered?",
                "answer": trust_principles_result,
                "status": determine_status("Are all three Trust Principles covered?", trust_principles_result)
            },
            {
                "question": "Does the SOC 2 Type 2 Report cover an audit period of at least 9 months?",
                "answer": audit_period_result,
                "status": determine_status("Does the SOC 2 Type 2 Report cover an audit period of at least 9 months?", audit_period_result)
            },
            {
                "question": "Are there any observations in the independent auditor’s opinion that signify the report is invalid?",
                "answer": invalid_observations_result,
                "status": determine_status("Are there invalid observations?", invalid_observations_result)
            },
            {
                "question": "Is the SOC 2 Type 2 Report a qualified report?",
                "answer": report_qualified_result,
                "status": determine_status("Is the SOC 2 Type 2 Report a qualified report?", report_qualified_result)
            }
        ]
        scope_of_services_result = describe_scope_of_services(df_temp, model, index, top_k=3)
        qualifiers.append({
            "question": "Please detail the scope of services within the SOC 2 Type 2 Report.",
            "answer": scope_of_services_result,
            "status": "N/A"
        })
        overall_viability = "Pass"
        for q in qualifiers:
            if q["status"] == "Fail":
                overall_viability = "Fail"
                break
        return jsonify({
            "qualifiers": qualifiers,
            "overall_viability": overall_viability
        }), 200
    except Exception as e:
        logging.error(f"Error in /initial_qualifier_check: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 400

# -------------------------------------------------------
# NEW ENDPOINT: /detect_control_ids (MODIFIED)
# -------------------------------------------------------
@app.route('/detect_control_ids', methods=['POST'])
def detect_control_ids_endpoint():
    logging.info("Received request to /detect_control_ids endpoint.")
    try:
        pdf_file = request.files.get('pdf_file')
        if not pdf_file:
            raise ValueError("Missing required file: pdf_file")
        filename_pdf = secure_filename(pdf_file.filename)
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename_pdf)
        pdf_file.save(pdf_path)
        logging.info(f"PDF file saved to {pdf_path}")
        repeating_patterns = identify_control_ids(pdf_path)
        regex_to_cids = {}
        for pattern_dict in repeating_patterns:
            regex = pattern_dict.get("Regex Pattern")
            cid = pattern_dict.get("Example Control ID")
            if regex and cid:
                regex_to_cids.setdefault(regex, []).append(cid)
        all_control_id_pages = detect_control_id_pages(pdf_path, regex_to_cids)
        section_iv_page = find_section_iv_page(pdf_path)
        logging.info(f"Section IV heading detected on page: {section_iv_page}")
        filtered_control_id_pages = {}
        if section_iv_page is not None:
            for cid, pages in all_control_id_pages.items():
                filtered_pages = [p for p in pages if p > section_iv_page]
                if filtered_pages:
                    filtered_control_id_pages[cid] = filtered_pages
                else:
                    logging.info(f"Control ID '{cid}' discarded (appears only before Section IV or not at all).")
        else:
            logging.info("No Section IV heading found; discarding all control IDs.")
            filtered_control_id_pages = {}
        return jsonify({
            "repeating_patterns": repeating_patterns,
            "control_id_pages": filtered_control_id_pages
        }), 200
    except Exception as e:
        logging.error(f"Error in /detect_control_ids endpoint: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 400

# -------------------------------------------------------
# UPDATED Endpoint: /process_all
# -------------------------------------------------------
@app.route('/process_all', methods=['POST'])
def process_all():
    logging.info("Received request to /process_all endpoint.")
    try:
        pdf_file = request.files.get('pdf_file')
        excel_file = request.files.get('excel_file')
        if not pdf_file or not excel_file:
            raise ValueError("Missing required files: pdf_file or excel_file")
        control_id = request.form.get('control_id', '').strip()
        if not control_id:
            logging.error("No Control IDs provided in the request.")
            return jsonify({"error": "No Control IDs were provided."}), 400
        filename_pdf = secure_filename(pdf_file.filename)
        filename_excel = secure_filename(excel_file.filename)
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename_pdf)
        pdf_file.save(pdf_path)
        logging.info(f"PDF file saved to {pdf_path}")
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], filename_excel)
        excel_file.save(excel_path)
        logging.info(f"Excel file saved to {excel_path}")
        repeating_patterns = identify_control_ids(pdf_path)
        regex_to_cids = {}
        for pattern_dict in repeating_patterns:
            regex = pattern_dict.get("Regex Pattern")
            cid = pattern_dict.get("Example Control ID")
            if regex and cid and cid in [c.strip() for c in control_id.split(',') if c.strip()]:
                regex_to_cids.setdefault(regex, []).append(cid)
        control_ids_order = [cid.strip() for cid in control_id.split(',') if cid.strip()]
        logging.info(f"Control IDs order for page range determination: {control_ids_order}")
        if not regex_to_cids:
            raise ValueError("No matching Control IDs found in the document based on the selected Control IDs.")
        control_id_pages = detect_control_id_pages(pdf_path, regex_to_cids)
        start_page, end_page = determine_page_range(control_id_pages, regex_to_cids)
        logging.info(f"Determined page range based on Control IDs: Start Page = {start_page}, End Page = {end_page}")
        task_id = str(uuid.uuid4())
        progress_data[task_id] = {
            'progress': 0.0,
            'status': 'Starting task...',
            'download_url': None,
            'error': None,
            'start_time': time.time(),
            'eta': None,
            'num_controls': 0,
            'cancelled': False
        }
        soc_report_filename = filename_pdf
        framework_filename = filename_excel
        thread = threading.Thread(
            target=background_process,
            args=(task_id, pdf_path, excel_path, start_page, end_page, control_id,
                  soc_report_filename, framework_filename)
        )
        thread.start()
        return jsonify({"message": "Processing started", "task_id": task_id}), 202
    except Exception as e:
        logging.error(f"Error in /process_all endpoint: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 400

@app.route('/progress/<task_id>', methods=['GET'])
def sse_progress(task_id):
    def generate():
        while True:
            task_info = progress_data.get(task_id, None)
            if task_info is None:
                yield f"data: {json.dumps({'error': 'Invalid task ID'})}\n\n"
                break
            progress = task_info['progress']
            status = task_info['status']
            eta = task_info.get('eta', None)
            error = task_info.get('error', None)
            download_url = task_info.get('download_url', None)
            cancelled = task_info.get('cancelled', False)
            data = {
                'progress': round(progress, 2),
                'status': status,
                'eta': round(eta, 2) if eta is not None else eta
            }
            if (download_url):
                data['download_url'] = download_url
            if (error):
                data['error'] = error
            if (cancelled):
                data['cancelled'] = True
            yield f"data: {json.dumps(data)}\n\n"
            if progress >= 100 or error or cancelled:
                break
            time.sleep(1)
    response = Response(generate(), mimetype='text/event-stream')
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['Connection'] = 'keep-alive'
    response.headers['X-Accel-Buffering'] = 'no'
    return response

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    logging.info(f"Received request to download file: {filename}")
    try:
        filename = secure_filename(filename)
        directory = app.config['EXCEL_FOLDER']
        file_path = os.path.join(directory, filename)
        if not os.path.exists(file_path):
            logging.error(f"File not found: {filename}")
            return jsonify({"error": "File not found"}), 404
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True
        )
    except Exception as e:
        logging.error(f"Error serving file {filename}: {e}", exc_info=True)
        return jsonify({"error": "File could not be served"}), 500

@app.route('/cancel_task/<task_id>', methods=['POST'])
def cancel_task(task_id):
    task_info = progress_data.get(task_id)
    if not task_info:
        logging.warning(f"Attempted to cancel invalid task_id: {task_id}.")
        return jsonify({"error": "Invalid task ID"}), 404
    if task_info.get('cancelled'):
        logging.info(f"Task {task_id} is already cancelled.")
        return jsonify({"message": f"Task {task_id} is already cancelled."}), 200
    task_info['cancelled'] = True
    task_info['status'] = "Cancelled"
    task_info['progress'] = 0
    task_info['eta'] = 0
    task_info['error'] = "Task was cancelled by the user."
    logging.info(f"Task {task_id} marked as cancelled by the user.")
    return jsonify({"message": f"Task {task_id} cancelled."}), 200

if __name__ == "__main__":
    logging.info("Starting the Flask application on port 5000.")
    app.run(host="0.0.0.0", port=5000, threaded=True, debug=False)
